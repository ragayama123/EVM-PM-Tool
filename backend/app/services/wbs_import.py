"""WBSインポート/エクスポートサービス"""

from datetime import datetime, date
from io import BytesIO
from typing import List, Optional, Tuple, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.task import Task
from app.models.member import Member


# タスク種別マッピング
TASK_TYPES = {
    '要件定義': 'requirements',
    '外部設計': 'external_design',
    '詳細設計': 'detailed_design',
    'PG': 'pg',
    'UT': 'ut',
    'CI': 'ci',
    'IT': 'it',
    'ST': 'st',
    '本番化': 'release',
}

TASK_TYPES_REVERSE = {v: k for k, v in TASK_TYPES.items()}


class WBSImportError:
    """インポートエラー情報"""
    def __init__(self, row: int, message: str):
        self.row = row
        self.message = message

    def to_dict(self):
        return {
            "row": self.row,
            "message": self.message
        }


class WBSImportTask:
    """インポート対象タスク情報"""
    def __init__(
        self,
        row: int,
        wbs_number: str,
        name: str,
        task_type: Optional[str] = None,
        planned_hours: float = 0,
        planned_start_date: Optional[date] = None,
        planned_end_date: Optional[date] = None,
        assigned_member_name: Optional[str] = None,
        description: Optional[str] = None,
        is_milestone: bool = False,
        predecessor_wbs: Optional[str] = None,
    ):
        self.row = row
        self.wbs_number = wbs_number
        self.name = name
        self.task_type = task_type
        self.planned_hours = planned_hours
        self.planned_start_date = planned_start_date
        self.planned_end_date = planned_end_date
        self.assigned_member_name = assigned_member_name
        self.description = description
        self.is_milestone = is_milestone
        self.predecessor_wbs = predecessor_wbs

        # 解決後の値
        self.assigned_member_id: Optional[int] = None
        self.predecessor_id: Optional[int] = None

    def to_dict(self):
        return {
            "row": self.row,
            "wbs_number": self.wbs_number,
            "name": self.name,
            "task_type": self.task_type,
            "task_type_label": TASK_TYPES_REVERSE.get(self.task_type, "") if self.task_type else "",
            "planned_hours": self.planned_hours,
            "planned_start_date": self.planned_start_date.isoformat() if self.planned_start_date else None,
            "planned_end_date": self.planned_end_date.isoformat() if self.planned_end_date else None,
            "assigned_member_name": self.assigned_member_name,
            "description": self.description,
            "is_milestone": self.is_milestone,
            "predecessor_wbs": self.predecessor_wbs,
        }


class WBSImportService:
    """WBSインポート/エクスポートサービス"""

    def __init__(self, db: Session, project_id: int):
        self.db = db
        self.project_id = project_id

    def generate_template(self) -> BytesIO:
        """Excelテンプレートを生成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "WBS"

        # ヘッダースタイル
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ヘッダー設定
        headers = [
            ("A", "WBS番号", 12),
            ("B", "タスク名", 30),
            ("C", "タスク種別", 15),
            ("D", "予定工数(h)", 12),
            ("E", "予定開始日", 15),
            ("F", "予定終了日", 15),
            ("G", "担当者", 15),
            ("H", "先行タスク", 12),
            ("I", "説明", 40),
            ("J", "固定日付", 10),
        ]

        for col, header, width in headers:
            cell = ws[f"{col}1"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[col].width = width

        # サンプルデータ（入力例として表示）
        # WBS番号, タスク名, タスク種別, 予定工数, 予定開始日, 予定終了日, 担当者, 先行タスク, 説明, 固定日付
        sample_data = [
            ("1", "【サンプル】要件ヒアリング", "要件定義", 16, "", "", "", "", "※サンプルデータは削除してください", ""),
            ("2", "【サンプル】要件定義書作成", "要件定義", 24, "", "", "", "1", "先行タスクはWBS番号で指定", ""),
            ("3", "【サンプル】外部設計", "外部設計", 40, "", "", "", "2", "", ""),
            ("4", "【サンプル】詳細設計", "詳細設計", 40, "", "", "", "3", "", ""),
            ("5", "【サンプル】プログラミング", "PG", 80, "", "", "", "4", "", ""),
        ]

        for row_idx, data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(data):
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                cell.border = thin_border

        # ドロップダウンリスト: タスク種別
        task_type_validation = DataValidation(
            type="list",
            formula1='"' + ','.join(TASK_TYPES.keys()) + '"',
            allow_blank=True
        )
        task_type_validation.error = "リストから選択してください"
        task_type_validation.errorTitle = "無効な入力"
        ws.add_data_validation(task_type_validation)
        task_type_validation.add(f"C2:C1000")

        # ドロップダウンリスト: 担当者
        members = self.db.query(Member).filter(Member.project_id == self.project_id).all()
        if members:
            member_names = ','.join([m.name for m in members])
            member_validation = DataValidation(
                type="list",
                formula1=f'"{member_names}"',
                allow_blank=True
            )
            member_validation.error = "リストから選択してください"
            member_validation.errorTitle = "無効な入力"
            ws.add_data_validation(member_validation)
            member_validation.add(f"G2:G1000")

        # ドロップダウンリスト: 固定日付
        milestone_validation = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True
        )
        ws.add_data_validation(milestone_validation)
        milestone_validation.add(f"J2:J1000")

        # 使い方シート
        ws_help = wb.create_sheet("使い方")
        help_content = [
            ["ExcelからのWBSインポート 使い方"],
            [""],
            ["■ WBS番号について"],
            ["  - タスクを識別するためのユニークな番号です"],
            ["  - 例: 1, 2, 3, 10, 20"],
            ["  - 数字でなくても構いません（例: A, B, C）"],
            [""],
            ["■ 先行タスクについて"],
            ["  - このタスクの前に完了すべきタスクのWBS番号を入力します"],
            ["  - 例: タスク「3」の先行タスクが「2」なら、H列に「2」と入力"],
            [""],
            ["■ 必須項目"],
            ["  - WBS番号: 必須"],
            ["  - タスク名: 必須"],
            [""],
            ["■ 日付形式"],
            ["  - YYYY-MM-DD形式で入力してください"],
            ["  - 例: 2025-01-15"],
            [""],
            ["■ 固定日付"],
            ["  - TRUEを設定すると、リスケジュール対象外になります"],
            [""],
            ["■ インポート時の注意"],
            ["  - インポートすると、既存のタスクは全て削除されます"],
            ["  - プレビュー画面で内容を確認してから実行してください"],
        ]
        for row_idx, row_data in enumerate(help_content, start=1):
            ws_help.cell(row=row_idx, column=1, value=row_data[0] if row_data else "")
        ws_help.column_dimensions["A"].width = 60

        # バイナリとして出力
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def parse_excel(self, file_content: bytes) -> Tuple[List[WBSImportTask], List[WBSImportError]]:
        """Excelファイルをパース"""
        from openpyxl import load_workbook

        errors: List[WBSImportError] = []
        tasks: List[WBSImportTask] = []

        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
        except Exception as e:
            errors.append(WBSImportError(0, f"Excelファイルの読み込みに失敗しました: {str(e)}"))
            return tasks, errors

        # WBSシートを探す
        ws = None
        if "WBS" in wb.sheetnames:
            ws = wb["WBS"]
        else:
            ws = wb.active

        # ヘッダー行をスキップしてデータ行を処理
        wbs_numbers_seen: Dict[str, int] = {}  # WBS番号 -> 行番号

        for row_idx in range(2, ws.max_row + 1):
            # 空行チェック
            wbs_number = ws.cell(row=row_idx, column=1).value
            name = ws.cell(row=row_idx, column=2).value

            if not wbs_number and not name:
                continue  # 空行はスキップ

            # WBS番号必須チェック
            if not wbs_number:
                errors.append(WBSImportError(row_idx, "WBS番号は必須です"))
                continue

            wbs_number = str(wbs_number).strip()

            # WBS番号重複チェック
            if wbs_number in wbs_numbers_seen:
                errors.append(WBSImportError(
                    row_idx,
                    f"WBS番号「{wbs_number}」が行{wbs_numbers_seen[wbs_number]}と重複しています"
                ))
                continue
            wbs_numbers_seen[wbs_number] = row_idx

            # タスク名必須チェック
            if not name:
                errors.append(WBSImportError(row_idx, "タスク名は必須です"))
                continue

            name = str(name).strip()

            # タスク種別
            task_type_raw = ws.cell(row=row_idx, column=3).value
            task_type = None
            if task_type_raw:
                task_type_str = str(task_type_raw).strip()
                if task_type_str in TASK_TYPES:
                    task_type = TASK_TYPES[task_type_str]
                elif task_type_str in TASK_TYPES.values():
                    task_type = task_type_str

            # 予定工数
            planned_hours_raw = ws.cell(row=row_idx, column=4).value
            planned_hours = 0.0
            if planned_hours_raw:
                try:
                    planned_hours = float(planned_hours_raw)
                except ValueError:
                    errors.append(WBSImportError(row_idx, f"予定工数「{planned_hours_raw}」は数値で入力してください"))

            # 予定開始日
            planned_start_raw = ws.cell(row=row_idx, column=5).value
            planned_start_date = None
            if planned_start_raw:
                planned_start_date = self._parse_date(planned_start_raw, row_idx, "予定開始日", errors)

            # 予定終了日
            planned_end_raw = ws.cell(row=row_idx, column=6).value
            planned_end_date = None
            if planned_end_raw:
                planned_end_date = self._parse_date(planned_end_raw, row_idx, "予定終了日", errors)

            # 日付の整合性チェック
            if planned_start_date and planned_end_date and planned_start_date > planned_end_date:
                errors.append(WBSImportError(row_idx, "予定開始日は予定終了日以前にしてください"))

            # 担当者
            assigned_member_name = ws.cell(row=row_idx, column=7).value
            if assigned_member_name:
                assigned_member_name = str(assigned_member_name).strip()

            # 先行タスク（WBS番号）
            predecessor_wbs = ws.cell(row=row_idx, column=8).value
            if predecessor_wbs:
                predecessor_wbs = str(predecessor_wbs).strip()

            # 説明
            description = ws.cell(row=row_idx, column=9).value
            if description:
                description = str(description).strip()

            # 固定日付
            is_milestone_raw = ws.cell(row=row_idx, column=10).value
            is_milestone = False
            if is_milestone_raw:
                is_milestone_str = str(is_milestone_raw).strip().upper()
                is_milestone = is_milestone_str in ("TRUE", "1", "はい", "YES")

            task = WBSImportTask(
                row=row_idx,
                wbs_number=wbs_number,
                name=name,
                task_type=task_type,
                planned_hours=planned_hours,
                planned_start_date=planned_start_date,
                planned_end_date=planned_end_date,
                assigned_member_name=assigned_member_name,
                description=description,
                is_milestone=is_milestone,
                predecessor_wbs=predecessor_wbs,
            )
            tasks.append(task)

        # 先行タスクWBS番号の存在チェック
        wbs_set = set(wbs_numbers_seen.keys())
        for task in tasks:
            if task.predecessor_wbs and task.predecessor_wbs not in wbs_set:
                errors.append(WBSImportError(
                    task.row,
                    f"先行タスク「{task.predecessor_wbs}」が見つかりません"
                ))

        return tasks, errors

    def _parse_date(
        self,
        value,
        row: int,
        field_name: str,
        errors: List[WBSImportError]
    ) -> Optional[date]:
        """日付をパース"""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        try:
            return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
        except ValueError:
            try:
                return datetime.strptime(str(value).strip(), "%Y/%m/%d").date()
            except ValueError:
                errors.append(WBSImportError(
                    row,
                    f"{field_name}「{value}」の日付形式が不正です（YYYY-MM-DD）"
                ))
                return None

    def resolve_references(
        self,
        tasks: List[WBSImportTask]
    ) -> Tuple[List[WBSImportTask], List[WBSImportError]]:
        """担当者名からIDへの解決"""
        errors: List[WBSImportError] = []

        # メンバー名→IDマッピングを作成
        members = self.db.query(Member).filter(Member.project_id == self.project_id).all()
        member_map = {m.name: m.id for m in members}

        for task in tasks:
            # 担当者の解決
            if task.assigned_member_name:
                if task.assigned_member_name in member_map:
                    task.assigned_member_id = member_map[task.assigned_member_name]
                else:
                    errors.append(WBSImportError(
                        task.row,
                        f"担当者「{task.assigned_member_name}」が見つかりません"
                    ))

        return tasks, errors

    def preview(self, file_content: bytes) -> dict:
        """インポートプレビュー"""
        tasks, parse_errors = self.parse_excel(file_content)

        if parse_errors:
            return {
                "success": False,
                "errors": [e.to_dict() for e in parse_errors],
                "tasks": [],
                "total_count": 0,
            }

        tasks, resolve_errors = self.resolve_references(tasks)

        all_errors = parse_errors + resolve_errors

        return {
            "success": len(all_errors) == 0,
            "errors": [e.to_dict() for e in all_errors],
            "tasks": [t.to_dict() for t in tasks],
            "total_count": len(tasks),
        }

    def execute_import(self, file_content: bytes) -> dict:
        """インポート実行（既存タスク削除→新規作成）"""
        tasks, parse_errors = self.parse_excel(file_content)

        if parse_errors:
            return {
                "success": False,
                "message": "バリデーションエラーがあります",
                "errors": [e.to_dict() for e in parse_errors],
                "imported_count": 0,
            }

        tasks, resolve_errors = self.resolve_references(tasks)

        if resolve_errors:
            return {
                "success": False,
                "message": "参照解決エラーがあります",
                "errors": [e.to_dict() for e in resolve_errors],
                "imported_count": 0,
            }

        # 既存タスクを全削除
        self.db.query(Task).filter(Task.project_id == self.project_id).delete()

        # 新規タスクを作成（まず先行タスクなしで作成）
        wbs_to_db_id: Dict[str, int] = {}
        created_tasks: List[Task] = []

        for task in tasks:
            db_task = Task(
                project_id=self.project_id,
                parent_id=None,  # 親子関係は使用しない
                predecessor_id=None,  # 後で設定
                name=task.name,
                description=task.description,
                task_type=task.task_type,
                planned_hours=task.planned_hours,
                planned_start_date=task.planned_start_date,
                planned_end_date=task.planned_end_date,
                assigned_member_id=task.assigned_member_id,
                is_milestone=task.is_milestone,
                progress=0,
                actual_hours=0,
                hourly_rate=5000,
            )
            self.db.add(db_task)
            self.db.flush()  # IDを取得するためにflush

            wbs_to_db_id[task.wbs_number] = db_task.id
            created_tasks.append(db_task)

        # 先行タスクIDを設定
        for i, task in enumerate(tasks):
            if task.predecessor_wbs and task.predecessor_wbs in wbs_to_db_id:
                created_tasks[i].predecessor_id = wbs_to_db_id[task.predecessor_wbs]

        self.db.commit()

        return {
            "success": True,
            "message": f"{len(tasks)}件のタスクをインポートしました",
            "errors": [],
            "imported_count": len(tasks),
        }
